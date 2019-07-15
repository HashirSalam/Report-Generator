import os

import requests
import pandas as pd
from io import StringIO
import time
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.writer.excel import ExcelWriter

yearsnow = time.strftime("%Y")
monthsnow = time.strftime("%m")
daynow = time.strftime("%d")
date = daynow+'-'+monthsnow+'-'+yearsnow
epoch_ago = time.time()-604800
yearsago = (time.strftime('%Y', time.localtime(epoch_ago)))
agomonths = (time.strftime('%m', time.localtime(epoch_ago)))
agodays = (time.strftime('%d', time.localtime(epoch_ago)))
date_ago = agodays+'-'+agomonths+'-'+yearsago

filename = (date_ago + ' to ' + date)
print(filename)

wb2 = load_workbook(r'.\Summary.xlsx')
wb2.create_sheet(filename)

ws1 = wb2.get_sheet_by_name(filename)
row = 1
ws1.cell(row=row, column=1).value = "Client"
ws1.cell(row=row, column=2).value = "Amount"
# ws1['A2'] = "SomeValue1"
# ws1['A2'] = "SomeValue1"


# data=[('Account','Amount')]
# sheet.append(['Client','Amount'])

path = ('C:/Users/User/Dropbox/PYCharms/PhoneDivert Report/Reports/' + date_ago + ' to ' + date)
os.mkdir(path)


class PhoneDivert():

    # yearsnow = time.strftime("%Y")
    # monthsnow = time.strftime("%m")
    # daynow = time.strftime("%d")
    # epoch_ago = time.time()-604800
    # yearsago = str((time.strftime('%Y', time.localtime(epoch_ago))))
    # agomonths = (time.strftime('%m', time.localtime(epoch_ago)))
    # agodays = (time.strftime('%d', time.localtime(epoch_ago)))
    #
    # print(daynow, daysago)
    def __init__(self, Client, Hunt_ID, Price):
        self.Client = Client
        self.Hunt_ID = Hunt_ID
        self.Price = Price

    def GetSeshRequest(self):

        headers = {
            'Content-type': 'application/x-www-form-urlencoded'
        }

        ses = requests.Session()
        r2 = ses.post('https://www.phonedivert.co.uk/login.php')
        r3 = ses.get('https://www.phonedivert.co.uk/statistics?number=&searchtype=huntgroup&huntgroup='+self.Hunt_ID+'&descr=&daterange=custom&fromdate='+agodays+'%2F'+agomonths+'%2F'+yearsago+'&todate='+daynow+'%2F'+monthsnow+'%2F'+yearsnow+'&stats=csv')
        return r3

    def DataFrameRequest(self):
        r3 = self.GetSeshRequest()
        decodecsv = r3.content.decode("UTF-8")
        # reader = csv.reader(decodecsv)
        file = StringIO(decodecsv)
        df = pd.read_csv(file, sep=",", header=None, names=['Date', 'Time', 'Calling Number', 'Called From', 'Called Number', 'Destination','Tme on Phone'])
        print(df)
        dedup = df.drop_duplicates(subset='Calling Number')
        print(dedup)
        count_row = dedup.shape[0]
        print(self.Price)
        print(count_row)
        amount = count_row * self.Price
        export_csv = dedup.to_csv(path+'/'+client+'.csv',index=None)
        row1 = row + 1
        row1 = int(row1)
        ws1.cell(row=row1, column=1).value = client
        ws1.cell(row=row1, column=2).value = amount
        return dedup
        return count_row



df = pd.read_excel(r'C:\Users\User\Dropbox\PYCharms\PhoneDivert Report\Clients.xlsx')
print(df)


for index, row in df.iterrows():
    client = str(row["Client"])
    Hunt_ID = str(row["Hunt ID"])
    Price = row["Price"]
    link = PhoneDivert(Client=client, Hunt_ID=Hunt_ID, Price=Price)
    PhoneDivert.DataFrameRequest(link)
    print(client, Hunt_ID, Price)

wb2.save('Summary.xlsx')