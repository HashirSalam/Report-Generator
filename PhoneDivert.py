import os
import sys
import glob
import requests
import pandas as pd
from io import StringIO
import time
import csv
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.writer.excel import ExcelWriter
from itertools import zip_longest
import xlsxwriter

pd.options.mode.chained_assignment = None # Removes warning for copied dataframe

def genrateClientReport():
    yearsnow = time.strftime("%Y")
    monthsnow = time.strftime("%m")
    daynow = time.strftime("%d")
    date = daynow+'-'+monthsnow+'-'+yearsnow
    epoch_ago = time.time()-604800
    yearsago = (time.strftime('%Y', time.localtime(epoch_ago)))
    agomonths = (time.strftime('%m', time.localtime(epoch_ago)))
    agodays = (time.strftime('%d', time.localtime(epoch_ago)))
    date_ago = agodays+'-'+agomonths+'-'+yearsago

    filename = (date_ago + ' to ' + date)
   

    #Loading data
    #data = input("Enter input filename (example: input.csv) : ") 
    data = pd.read_csv("input.csv")
    #separating witheld information
    reservedRecords = data.loc[data['Caller'] == ("withheld")]

    # dropping ALL duplicte values 
    data.drop_duplicates(subset ="Caller", keep = False, inplace = True)

    #Add 0s for numbers
    data['Caller'] = '0' + data['Caller'].astype(str)
    data['Called'] = '0' + data['Called'].astype(str)
    data['Destination'] = '0' + data['Destination'].astype(str)

    reservedRecords['Called'] = '0' + reservedRecords['Called'].astype(str)
    reservedRecords['Destination'] = '0' + reservedRecords['Destination'].astype(str)
    #To string
    data['Caller'] = '"'+data['Caller']+'"'
    data['Called'] = '"'+data['Called']+'"'
    data['Destination'] = '"'+data['Destination']+'"'

    reservedRecords['Called'] = '"'+reservedRecords['Called']+'"'
    reservedRecords['Destination'] = '"'+reservedRecords['Destination']+'"'

    #Concat with rest 
    frames = [data, reservedRecords]
    result = pd.concat(frames)
    
    #Sort according to date
    result = result.sort_values(by=['Date'])

    #Calculate total duration
    Duration = data['Duration'].values.tolist()
    hours =[]
    minutes = []
    for d in Duration:
        hour, minute = d.split(':')
        hours.append(hour)
        minutes.append(minute)
    totalHours = sum(map(int, hours))
    totalMinutes = sum(map(int, minutes))    
    totalHours = totalHours + (totalMinutes/60) 
    totalMinutes = totalMinutes % 60
    totalDuration = str(int(totalHours)) +":"+ str(int(totalMinutes))
    
    totalDuration = '"'+totalDuration+'"'
    #ClientName = input("Enter client's name : ") 
    ClientName = "Skip Hire"

    #Calculation
    rowCount = result.shape[0]
    TotalInquries = rowCount * 0.75
    
    #Getting price according to client
    Clients = pd.read_excel('Clients.xlsx', index_col=0) 
    Clients = Clients[['Client', 'Price']]
    Price = Clients.loc[Clients['Client'] == (ClientName)]
    Price = Price['Price'].values[0]
    TotalPrice = Price *TotalInquries

    # Create caclulation frame on CSV top
    df1 = pd.DataFrame({
    'Date': [ClientName, 'Date Range', '', '','',''],
    'Time': ['',filename, str(rowCount)+' -25%', '', '', ''],
    'Caller': ['', '', 'Total enquires', 'Total Time on Phone (minutes)','',''],
    'Location': ['', '', str(TotalInquries)+' Invoiced', str(totalDuration),'',''],
    'Called': ['', '', 'x'+str(Price), '','',''],
    'Destination': ['', '', "£"+str(TotalPrice), '','',''],
    'Duration': ['', '', '', '','','']}
    ,index=[0, 1, 2, 3,4,5])
    result = df1.append(result)
    #print (result)
    #Create Report
    path = "./Reports/" + date_ago + " to " + date
    #Create Directory if doesnt exit
    if not os.path.exists(path):
        os.makedirs(path)
    #Create file in directory
    filename = path+"/" + ClientName + ".csv"
    result.to_csv(filename,encoding='utf-8-sig', index=False, header=False)
    print("Report generated in " + filename)

##########################
def genrateSummaryReport():
    
    folders = [f for f in glob.glob('.\\Reports\\' + "**/", recursive=False)]
    choices = []
    for f in folders:
        choices.append(f)

    writer = pd.ExcelWriter("test.xlsx", engine = 'xlsxwriter')
    #loop over each folder in Reports 
    for interval in choices: 
        dateInterval = interval
        path = r'%s' % dateInterval
        #path = r".\Reports\"+ dateInterval # use your path
        all_files = glob.glob(path + "/*.csv")
        li = []

        #Extracting just 2 lines from the CSVs
        for filename in all_files:
            df = pd.read_csv(filename, nrows=2)
            li.append(df)

        frame = pd.concat(li, axis=0, ignore_index=True,sort=True)
        #print (frame)
        #For those who might want, for example, every fifth row, but starting at the 2nd row it would be df.iloc[1::5, :]
        #df.iloc[:, n]   to access the column at the nth position
        clients = list(frame.columns.values)
        clients =  [x for x in clients if "Unnamed:" not in x]

        prices = frame.iloc[1::2, 7::7].values.tolist()
        prices = [val for sublist in prices for val in sublist]
        prices = [s.replace('£', '') for s in prices]

        enquires = frame.iloc[1::2, 5::5].values.tolist()
        enquires = [val for sublist in enquires for val in sublist]
        enquires = [s.replace(' Invoiced', '') for s in enquires]
       
        #merge all information        
        po = zip(clients,enquires, prices)

        sheetName= dateInterval.replace('.\\Reports',"")
        sheetName= sheetName.replace("\\","")    
        print(sheetName)
        df = pd.DataFrame(po)
        #Updating header information
        df.columns = ['Clients', 'Enquires', 'Prices']
        #print(df)
        df.to_excel(writer, sheet_name = sheetName,index=False,header=True)
     
    writer.save()
    writer.close()
    print("Summary report generated successfully")

###########################
if __name__== "__main__":
  genrateClientReport()
  genrateSummaryReport()
##################################################################################################################################
#os.mkdir(path)
# wb2 = load_workbook(r'.\Summary.xlsx')
# wb2.create_sheet(filename)
# ws1 = wb2.get_sheet_by_name(filename)
# row = 1
# ws1.cell(row=row, column=1).value = "Client"
# ws1.cell(row=row, column=2).value = "Amount"
# # ws1['A2'] = "SomeValue1"
# # ws1['A2'] = "SomeValue1"
# # data=[('Account','Amount')]
# # sheet.append(['Client','Amount'])
# path = "./Reports/" + date_ago + " to " + date
# os.mkdir(path)
# class PhoneDivert():

#     # yearsnow = time.strftime("%Y")
#     # monthsnow = time.strftime("%m")
#     # daynow = time.strftime("%d")
#     # epoch_ago = time.time()-604800
#     # yearsago = str((time.strftime('%Y', time.localtime(epoch_ago))))
#     # agomonths = (time.strftime('%m', time.localtime(epoch_ago)))
#     # agodays = (time.strftime('%d', time.localtime(epoch_ago)))
#     #
#     # print(daynow, daysago)
#     def __init__(self, Client, Hunt_ID, Price):
#         self.Client = Client
#         self.Hunt_ID = Hunt_ID
#         self.Price = Price

#     def GetSeshRequest(self):

#         headers = {
#             'Content-type': 'application/x-www-form-urlencoded'
#         }

#         ses = requests.Session()
#         r2 = ses.post('https://www.phonedivert.co.uk/login.php')
#         r3 = ses.get('https://www.phonedivert.co.uk/statistics?number=&searchtype=huntgroup&huntgroup='+self.Hunt_ID+'&descr=&daterange=custom&fromdate='+agodays+'%2F'+agomonths+'%2F'+yearsago+'&todate='+daynow+'%2F'+monthsnow+'%2F'+yearsnow+'&stats=csv')
#         return r3

#     def DataFrameRequest(self):
#         r3 = self.GetSeshRequest()
#         decodecsv = r3.content.decode("UTF-8")
#         # reader = csv.reader(decodecsv)
#         file = StringIO(decodecsv)
#         df = pd.read_csv(file, sep=",", header=None, names=['Date', 'Time', 'Calling Number', 'Called From', 'Called Number', 'Destination','Tme on Phone'])
#         print(df)
#         dedup = df.drop_duplicates(subset='Calling Number')
#         print(dedup)
#         count_row = dedup.shape[0]
#         print(self.Price)
#         print(count_row)
#         amount = count_row * self.Price
#         export_csv = dedup.to_csv(path+'/'+client+'.csv',index=None)
#         row1 = row + 1
#         row1 = int(row1)
#         ws1.cell(row=row1, column=1).value = client
#         ws1.cell(row=row1, column=2).value = amount
#         return dedup
#         return count_row



# df = pd.read_excel(r'C:\Users\User\Dropbox\PYCharms\PhoneDivert Report\Clients.xlsx')
# print(df)


# for index, row in df.iterrows():
#     client = str(row["Client"])
#     Hunt_ID = str(row["Hunt ID"])
#     Price = row["Price"]
#     link = PhoneDivert(Client=client, Hunt_ID=Hunt_ID, Price=Price)
#     PhoneDivert.DataFrameRequest(link)
#     print(client, Hunt_ID, Price)

# wb2.save('Summary.xlsx')